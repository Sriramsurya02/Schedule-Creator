import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
path = "D:\\Code\\Schedule\\schedule.xlsx"
wb = openpyxl.Workbook()    
ws = wb.active
ws.title = "Schedule"
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

start_time = datetime.time(9, 0)
end_time = datetime.time(17, 0)

ws.column_dimensions['A'].width = 5
for i in range(1, 15):
      if i % 2 == 0:
       ws.cell(row=1, column=i).value = days[(i//2)-1]
      else:
       ws.cell(row=1, column=i).value = ""
# Set column widths
column_widths = [10, 20]
for i in enumerate(range(1, 15), start=1):
   if i[0] % 2 == 0:
    ws.column_dimensions[get_column_letter(i[0])].width = column_widths[0]
   else:
    ws.column_dimensions[get_column_letter(i[0])].width = column_widths[1]

while start_time < end_time:
    row = (start_time.hour - 9) * 2 + (1 if start_time.minute >= 30 else 0) + 2
    ws.cell(row=row, column=1).value = start_time.strftime("%H:%M")
    start_time = (datetime.datetime.combine(datetime.date.today(), start_time) + datetime.timedelta(minutes=30)).time()

# Apply styles
header_font = Font(bold=True)
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Get user input for tasks
def get_tasks():
   tasks = {}
   ip = input("Enter tasks (day time task), or 'done' to finish: ")
   while ip.lower() != 'done':
      try:
         day, time, task = ip.split(maxsplit=2)
         if day in days:
            tasks[(day, time)] = task
         else:
            print("Invalid day. Please enter a valid day of the week.")
      except ValueError:
         print("Invalid input format. Please enter in the format: 'day time task'.")
      ip = input("Enter tasks (day time task), or 'done' to finish: ")
   return tasks

# Fill in the schedule with user input tasks
tasks = get_tasks()
for (day, time), task in tasks.items():
   col_start = days.index(day)*2 + 2
   row = (int(time.split(':')[0]) - 9) * 2 + (1 if int(time.split(':')[1]) >= 30 else 0) + 2
   ws.cell(row=row, column=col_start).value = task
   ws.cell(row=row, column=col_start).border = thin_border
wb.save(path)
print("Schedule saved to schedule.xlsx")
